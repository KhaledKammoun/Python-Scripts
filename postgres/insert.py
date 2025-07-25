import pymysql
import psycopg2

import os
from openpyxl import load_workbook
from dotenv import load_dotenv

seperator = "|"
def getAllRowsFromExcel(sheet):
    global seperator
    elements = []
    allRowsList = list(sheet.iter_rows(min_row=0, max_row=sheet.max_row))
    seperator = allRowsList[1][0].value
    # add rows value to the elements, distinct: id, name, description, niveau
    for row in allRowsList[1:]:
        element = [cell.value for cell in row][1:11]
        if (element[0]) : 
            elements.append([cell.value for cell in row][1:11])
    return elements

# get the excel file data
input_path = "excel_file.xlsx"
excelFile = load_workbook(input_path, data_only=True)
workSheet = excelFile.active
elements = getAllRowsFromExcel(workSheet)
print("Seperator : ", seperator)

load_dotenv()
# Connect to the MySQL database
# connection = pymysql.connect(
#     host=os.environ.get('MYSQL_HOST'),
#     user=os.environ.get('MYSQL_USER'),
#     password=os.environ.get('MYSQL_PASSWORD'),
#     database=os.environ.get('MYSQL_DATABASE')
# )

# Connect to PostgreSQL
connection = psycopg2.connect(
    host=os.environ.get('MYSQL_HOST'),
    user=os.environ.get('MYSQL_USER'),
    password=os.environ.get('MYSQL_PASSWORD'),
    database=os.environ.get('MYSQL_DATABASE'),
    port=os.environ.get('PORT')

)

user_id = 1
lot_id = 1
# Create a cursor object
cursor = connection.cursor()
sql = "INSERT INTO articles (lot_article_id, lot_id, article_content, parent_article, user_id, unite) VALUES (%s, %s, %s, %s, %s, %s) RETURNING article_id"
sql_descriptif = "INSERT INTO Descriptifs (article_id, descriptif_brd, descriptif_name, descriptif_origine, user_id) VALUES (%s, %s, %s, %s, %s) RETURNING descriptif_id"
sql_getArticleID = "select article_id from articles where lot_article_id = %s"
sql_select = "select * from articles"

sql_set_article = "update articles set descriptif_id = %s where article_id = %s"

i = 1


try:
    for element in elements :
        id, id1, id2, id3, name, _, level, unite, descriptif_name, descriptif = element
        id = "|".join(str(c) for c in id.split(seperator))

        print([id,id1, id2, id3, name, _, level, unite, descriptif_name, descriptif])
        # chapter
        if level == 1 :
            try:
                # Add chapter to the database without a parent id
                values = (id, lot_id, name, 379, user_id, unite)
                cursor.execute(sql, values)
                print(("*" * level) + str(values))

                result_article_id = cursor.fetchone()
                article_id = result_article_id[0]
                print(f"*Article ID: {article_id}")

                if descriptif:
                    try:
                        # Insert new descriptif
                        values_descriptif = (article_id, descriptif, descriptif_name, None, user_id)
                        cursor.execute(sql_descriptif, values_descriptif)
                        connection.commit()  # Commit the descriptif insertion

                        # Retrieve the last inserted descriptif ID
                        _result_descriptif = cursor.fetchone()
                        descriptif_id = _result_descriptif[0]
                        print(f"Descriptif ID: {descriptif_id}")

                        # Update article table with descriptif_id
                        values_set_article = (descriptif_id, article_id)
                        cursor.execute(sql_set_article, values_set_article)
                        connection.commit()  # Commit the article update
                        print("Article updated with Descriptif ID.")

                    except psycopg2.Error as e:
                        connection.rollback()  # Rollback in case of any exception
                        print(f"Error during descriptif insertion and update: {e}")

                else:
                    print("Descriptif is missing.")

            except psycopg2.Error as e:
                connection.rollback()  # Rollback in case of any exception
                print(f"Error during article insertion: {e}")

        else :
                # if the second id is 0, add descriptif to the chapter not the article

                if level == 2 :
                    # if the third id is 0, add article to the data base with descriptif if there .
                    id_parent = (id1 + "|00|00")
                    # sql = "INSERT INTO articles (lotArticleID, lotID, articleContent, parentArticle, userID) VALUES (%s, %s, %s, %s, %s)"
                    cursor.execute(sql_getArticleID, (id_parent,))
                    parentArticleID = cursor.fetchall()  # Fetch the first row
                    values = (id, lot_id, name, parentArticleID[-1], user_id, unite)
                    print(("*" * level) + str(values))
                    cursor.execute(sql, values)

                    # Retrieve the last inserted article ID
                    result_article_id = cursor.fetchone()
                    article_id = result_article_id[0]
                    print(f"**Article ID: {article_id}")

                    if descriptif:
                        try:
                            # Insert new descriptif
                            values_descriptif = (article_id, descriptif, descriptif_name, None, user_id)
                            cursor.execute(sql_descriptif, values_descriptif)
                            connection.commit()  # Commit the descriptif insertion

                            # Retrieve the last inserted descriptif ID
                            _result_descriptif = cursor.fetchone()
                            descriptif_id = _result_descriptif[0]
                            print(f"**Descriptif ID: {descriptif_id}")

                            # Update article table with descriptif_id
                            values_set_article = (descriptif_id, article_id)
                            cursor.execute(sql_set_article, values_set_article)
                            connection.commit()  # Commit the article update
                            print("Article updated with Descriptif ID.")

                        except psycopg2.Error as e:
                            connection.rollback()  # Rollback in case of any exception
                            print(f"Error: {e}")

                    else:
                        print("Descriptif is missing.")

                else :
                    # if the third id is not 0, add article to the data base with descriptif if there .
                    
                    id_parent = (id1 + "|"+id2+"|00")
                    # sql = "INSERT INTO articles (lotArticleID, lotID, articleContent, parentArticle, userID) VALUES (%s, %s, %s, %s, %s)"
                    cursor.execute(sql_getArticleID, (id_parent,))
                    parentArticleID = cursor.fetchall()  # Fetch the first row
                    if (parentArticleID is None) :
                         print(f"\n{id_parent} : is not exist\n")
                    values = (id, lot_id, name, parentArticleID[-1], user_id, unite)
                    print(("*" * level) + str(values))
                    cursor.execute(sql, values)
                    
                    # Retrieve the last inserted article ID
                    result_article_id = cursor.fetchone()
                    article_id = result_article_id[0]
                    print(f"***Article ID: {article_id}")

                    if descriptif:
                        try:
                            # Insert new descriptif
                            values_descriptif = (article_id, descriptif, descriptif_name, None, user_id)
                            cursor.execute(sql_descriptif, values_descriptif)
                            connection.commit()  # Commit the descriptif insertion

                            # Retrieve descriptif ID
                            _result_descriptif = cursor.fetchone()
                            descriptif_id = _result_descriptif[0]
                            print(f"***Descriptif ID: {descriptif_id}")

                            # Update article table with descriptif_id
                            values_set_article = (descriptif_id, article_id)
                            cursor.execute(sql_set_article, values_set_article)
                            connection.commit()  # Commit the article update
                            print("Article updated with Descriptif ID.")
                        except psycopg2.Error as e:
                            connection.rollback()  # Rollback in case of any exception
                            print(f"Error: {e}")

                    else:
                        print("Descriptif is missing.")


        i+=1
    # Commit Changes
    connection.commit()
    
except pymysql.Error as error:
    # Rollback in case of an error
    print("Error inserting data:", error)
    connection.rollback()

    # test the data base :

# try:
#     cursor.execute(sql_select)
#     result = cursor.fetchall()  # Fetch all rows from the result set
#     result = [element[4] for element in result]
#     dict_result = {}
#     for id in result:
#         if id in dict_result:
#             dict_result[id] += 1
#         else:
#             dict_result[id] = 1

#     for value in dict_result.keys() :
#         if dict_result[value] > 1 :
#             print(value, " :: ", dict_result[value])
# except pymysql.Error as error:
#     print("Error fetching data:", error)

# Close the cursor and connection
cursor.close()
connection.close()