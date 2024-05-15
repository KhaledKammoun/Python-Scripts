import json
from openpyxl import load_workbook

def getAllRowsFromExcel(sheet):
    elements = []
    allRowsList = list(sheet.iter_rows(min_row=2, max_row=sheet.max_row))
    
    # add rows value to the elements, distinct: id, name, description, niveau
    for row in allRowsList:
        elements.append([cell.value for cell in row][:6])
    return elements
input_path = "your_excel_file.xlsx"
excelFile = load_workbook(input_path, read_only=True)
workSheet = excelFile.active
elements = getAllRowsFromExcel(workSheet)

additional_data = {
    "Marge": None,
    "Q_manuelle": None,
    "Q_auto": None,
    
    "Prix_U_APS" : None,
    "Prix_U_APD" : None,
    "Prix_U_Dossier_C" : None,
    "Prix_U_Marche" : None,

    "Q_APS" : None,
    "Q_APD" : None,
    "Q_Dossier_C" : None,
    "Q_Marche" : None
}
def update(level, item) :
    data_var = item
    item = {}
    if (level == 1) :
        item["ID"] = int(data_var[0])
    elif level == 2 :
        item["ID"] = data_var[0] + "|" + data_var[1] + "|00"
    else :
        item["ID"] = data_var[0] + "|" + data_var[1] + "|" + data_var[2]
    item["Name"] = data_var[3]
    item["U"] = data_var[4]
    for key, value in additional_data.items():
        item[key] = value
    if (not data_var[5]) :
        item["Description"] = ""
    else :
        item["Description"] = [data_var[5]]
    item["Children"] = ""

    
    return item


data = dict()
data["Items"] = dict()
data["Items"]["Item"]  = []

for i in range(len(elements)) :
    # first id must start with 00 not 01
    if (elements[i][1] in [None, '']) :
        index_1 = int(elements[i][0])
        data["Items"]["Item"].append(update(1,elements[i]))
    elif elements[i][2] in ['00', None, ''] :
        
        # second id must start with 01 not 00
        index_1 = int(elements[i][0])
        children = data["Items"]["Item"][index_1]["Children"]
        if children == "":
            data["Items"]["Item"][index_1]["Children"] = {"Item" : [update(2,elements[i])]}
        else :
            data["Items"]["Item"][index_1]["Children"]["Item"].append(update(2,elements[i]))

    else :    
        # thrid id must start with 01 not 00
        index_1 = int(elements[i][0])
        last_level_2_element = data["Items"]["Item"][index_1]["Children"]["Item"]
        if last_level_2_element[len(last_level_2_element) -  1]["Children"] == "" :

            last_level_2_element[len(last_level_2_element) -  1]["Children"] = {"Item" : [update(3,elements[i])]}
        else :
            last_level_2_element[len(last_level_2_element) -  1]["Children"]["Item"].append(update(3, elements[i]))


# for element in elements :
#     print(element)
with open('your_file.json', 'w', encoding='utf-8') as file:
    json.dump(data, file, indent=2, ensure_ascii=False)
