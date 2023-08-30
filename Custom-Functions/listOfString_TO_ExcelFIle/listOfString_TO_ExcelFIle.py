import openpyxl
import os
def open_file(file_path):
    try:
        os.system(f"start {file_path}")
    except Exception as e:
        print(f"Error opening the file: {e}")

with open('input.txt') as f : # Paste in input.txt your values .

    float_list = [float(s) for s in f]

    # Prepare data for Excel with each float in its own list
    data_for_excel = [[value] for value in float_list]

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write the data to the Excel sheet
for col_idx, column in enumerate(data_for_excel, start=1):
    for row_idx, value in enumerate(column, start=1):
        sheet.cell(row=col_idx, column=row_idx, value=value) ## Values on The columns .
        # sheet.cell(row=row_idx, column=col_idx, value=value) ## Values on The rows .
file_path = "{}\{}".format(os.getcwd(),"output.xlsx")

# Save the Excel file, you can choose you own local save location .
workbook.save(file_path) # os.getcwd() : to guarantee the creation in your current python file .

# if u wanna open to open an Excel file after creating, call the open_file function .
open_file(file_path)