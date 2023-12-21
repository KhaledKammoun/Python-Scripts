from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import os
import xlwings as xw
import subprocess
def open_file(filePath) :
    try:
        subprocess.Popen(['start', 'excel', filePath], shell=True)
    except Exception as e:
        print(f"An error occurred: {e}")

def close_file(filePath) :
    try:
        book = xw.Book(filePath)
        book.close()
    except Exception as e:
        print(e)


def excelSheet_modulation(sheet) :

     # Convert the generator to a list for reversing
    rows_to_delete = list(sheet.iter_rows(min_row=1, max_row=sheet.max_row))
    
    
    # Delete empty rows
    for row in reversed(rows_to_delete):
        if all(cell.value is None for cell in row):
            sheet.delete_rows(row[0].row, amount=1)

    # Convert the generator to a list for reversing
    cols_to_delete = list(sheet.iter_cols(min_col=1, max_col=sheet.max_column))
    
    # Delete empty columns
    for col in reversed(cols_to_delete):
        if all(cell.value is None for cell in col):
            sheet.delete_cols(col[0].col_idx, amount=1)
    
    return sheet
# close_file('FamilyExcel.xlsx')
excelFile = load_workbook('FamilyExcel.xlsx')
workSheet = excelFile.active
# workSheet = excelSheet_modulation(workSheet)

class ExcelElementsClass :
    def __init__(self, id, name, description, niveau) :
        self.id = id
        self.name = name
        self.description = description
        self.niveau = niveau
    @staticmethod
    def getAllRowsFromExcel(sheet):
        elements = []
        allRowsList = list(sheet.iter_rows(min_row=2, max_row=sheet.max_row))
        
        # add rows value to the elements, distinct: id, name, description, niveau
        for row in allRowsList:
            elements.append(TreeNode(ExcelElementsClass(*[cell.value for cell in row][:4])))

        return elements

class TreeNode:
    def __init__(self, data):
        self.data = data
        self.children = []

    def add_child(self, child_node):
        self.children.append(child_node)

def print_tree(node, level=0, prefix="Root"):
    if level == 0:
        print(f"{prefix} - {node.data.name}")
    else:
        indent = " " * (level * 4)
        print(f"{indent}└── {node.data.niveau} - {node.data.name}")

    for child in node.children:
        print_tree(child, level + 1, f"{prefix}.{child.data.niveau}")

# contain all the excel rows with distinct value according to ('id', 'name', 'description', 'Niveau')
elements = ExcelElementsClass.getAllRowsFromExcel(workSheet)

def createTree(root) :
    queue = [root, elements[0]]
    root.add_child(elements[0])
    for i in range(1, len(elements)) :
        if elements[i - 1].data.niveau < elements[i].data.niveau :
            queue[len(queue) - 1].add_child(elements[i])
        else :
            while (elements[i].data.niveau <= queue[len(queue) - 1].data.niveau) :
                queue.pop()
            queue[len(queue) - 1].add_child(elements[i])
        queue.append(elements[i])
    return root
# Create a Tree
root = TreeNode(ExcelElementsClass('0', "Persons",None, 0))

createTree(root)
# Print the tree starting from the root
print_tree(root)
# excelFile.save('FamilyExcel.xlsx')
# open_file('FamilyExcel.xlsx')