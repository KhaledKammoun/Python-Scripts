import pymysql
import os
from openpyxl import load_workbook
from dotenv import load_dotenv
def getAllRowsFromExcel(sheet):
    elements = []
    allRowsList = list(sheet.iter_rows(min_row=0, max_row=sheet.max_row))
    
    # add rows value to the elements, distinct: id, name, description, niveau
    for row in allRowsList:
        element = [cell.value for cell in row][:6]
        if (element[0]) : 
            elements.append([cell.value for cell in row][:4])
    return elements

# get the excel file data
input_path = "excel_file.xlsx"
excelFile = load_workbook(input_path, read_only=True)
workSheet = excelFile.active
elements = getAllRowsFromExcel(workSheet)


load_dotenv()
# Connect to the MySQL database
connection = pymysql.connect(
    host=os.environ.get('MYSQL_HOST'),
    user=os.environ.get('MYSQL_USER'),
    password=os.environ.get('MYSQL_PASSWORD'),
    database=os.environ.get('MYSQL_DATABASE')
)
# Create a cursor object
cursor = connection.cursor()
sql = "INSERT INTO articles (lotArticleID, lotID, articleContent, parentArticle, userID) VALUES (%s, %s, %s, %s, %s)"
sql_descriptif = "INSERT INTO Descriptifs (articleID, descriptif_brd) VALUES (%s, %s)"
sql_getArticleID = "select articleID from articles where lotArticleID = %s"
sql_select = "select * from articles"

i = 1
try:
    for element in elements :
        id, name, descriptif, level = element
        # chapter

        if len(id) == 2 :
            # add chapter to the database without a parent id
            values = (id+"|00|00", 1, name, None, 1)
            # Execute the SQL query
            # cursor.execute(sql, values)
            # print("***1 : ", values)
            # print("{} : Data inserted successfully!".format(i))
        else :
            id1, id2, id3 = id.split('|')
            # if the second id is 0, add descriptif to the chapter not the article
            if (not int(id2) and descriptif) :
                # get articleID
                cursor.execute(sql_getArticleID, (id,))
                result = cursor.fetchone()  # Fetch the first row
                # print("ID : ", id)
                values_descriptif = (result[0], descriptif)
                cursor.execute(sql_descriptif, values_descriptif)

            elif not int(id3) :
                # if the third id is 0, add article to the data base with descriptif if there .
                # id_parent = (id1 + "|00|00")
                # sql = "INSERT INTO articles (lotArticleID, lotID, articleContent, parentArticle, userID) VALUES (%s, %s, %s, %s, %s)"
                # cursor.execute(sql_getArticleID, (id_parent,))
                # parentArticleID = cursor.fetchone()  # Fetch the first row
                # values = (id, 1, name, parentArticleID[0], 1)
                # cursor.execute(sql, values)

                if descriptif :
                    # print(values)
                    # get articleID
                    cursor.execute(sql_getArticleID, (id,))
                    result = cursor.fetchone()  # Fetch the first row
                    # print("ID : ", result[0])
                    values_descriptif = (result[0], descriptif)
                    cursor.execute(sql_descriptif, values_descriptif)
            else :
                # if the third id is not 0, add article to the data base with descriptif if there .
                
                # id_parent = (id1 + "|"+id2+"|00")
                # sql = "INSERT INTO articles (lotArticleID, lotID, articleContent, parentArticle, userID) VALUES (%s, %s, %s, %s, %s)"
                # cursor.execute(sql_getArticleID, (id_parent,))
                # parentArticleID = cursor.fetchone()  # Fetch the first row
                # values = (id, 1, name, parentArticleID[0], 1)
                # print(values)
                # cursor.execute(sql, values)
                if descriptif :
                    # print(values)
                    
                    # get articleID
                    cursor.execute(sql_getArticleID, (id,))
                    result = cursor.fetchone()  # Fetch the first row
                    # print("ID : ", result[0])
                    values_descriptif = (result[0], descriptif)
                    cursor.execute(sql_descriptif, values_descriptif)
        i+=1
    # Commit Changes
    connection.commit()
    
except pymysql.Error as error:
    # Rollback in case of an error
    print("Error inserting data:", error)
    connection.rollback()


# test the data base :

try:
    cursor.execute(sql_select)
    result = cursor.fetchall()  # Fetch all rows from the result set
    result = [element[4] for element in result]
    dict_result = {}
    for id in result:
        if id in dict_result:
            dict_result[id] += 1
        else:
            dict_result[id] = 1

    for value in dict_result.keys() :
        if dict_result[value] > 1 :
            print(value, " :: ", dict_result[value])
except pymysql.Error as error:
    print("Error fetching data:", error)

# Close the cursor and connection
cursor.close()
connection.close()